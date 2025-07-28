from labgenpackage.classes import Student, Group
from xlsxwriter.utility import xl_rowcol_to_cell
from gui.util import CopyAndRename
from pathlib import Path

import openpyxl.worksheet.worksheet
import openpyxl.worksheet
import openpyxl
import gui.settings as settings
import xlsxwriter, locale, re, logging

# custom Exception class
class BadWorkbook(Exception):
    pass

# -------------------------------------------------------------------------------
# excel functions

# Checks if loaded workbook is usable.
# Returns -> tuple[bool,bool] = (usable,type)
# usable: true = good workbook, false = bad workbook
# type: true = workbook from 'program', false = workbook from 'merlin'
def CheckForValidInputWorkbook(wb: openpyxl.Workbook, sh: openpyxl.worksheet.worksheet.Worksheet) -> str:
    try:
        if not len(wb.sheetnames) == 1 or \
           not sh.max_column == 6 or \
           not sh.cell(row = 1, column = 1).value == 'Prezime' or \
           not sh.cell(row = 1, column = 2).value == 'Ime' or \
           not (sh.cell(row = 1, column = 4).value == 'ID broj' or sh.cell(row = 1, column = 4).value == 'JMBAG'):
            raise BadWorkbook('Loaded input workbook is not apropriet.')
        
        if sh.cell(row = 1, column = 3).value == 'Email' and \
           sh.cell(row = 1, column = 5).value == 'Korisničko ime' and \
           sh.cell(row = 1, column = 6).value == 'Grupa':
            return 'program'
        elif sh.cell(row = 1, column = 3).value == 'Korisničko ime' and \
             sh.cell(row = 1, column = 5).value == 'Grupa' and \
             sh.cell(row = 1, column = 6).value == 'Odabir {$a}':
            return 'merlin'
        else:
            raise BadWorkbook('Loaded input workbook is not apropriet.')
        
    except BadWorkbook as e:
        e.add_note('Prilozena excel datoteka sa podatcima o ispunjenim grupama nije prikladna!')
        raise
    except Exception:
        Exception.add_note('Unexpected error with input wb validation!')
        raise

# Checks if loaded workbook is usable.
# Returns -> bool = usable
# usable: true = good workbook, false = bad workbook
def CheckForValidOldWorkbook(wb: openpyxl.Workbook, sh: openpyxl.worksheet.worksheet.Worksheet) -> str:
    try:
        if not (len(wb.sheetnames) == 4 or len(wb.sheetnames) == 1) or \
           not (sh.max_column == 15 or sh.max_column == 10) or \
           not sh.cell(row = 1, column = 1).value == 'Prezime' or \
           not sh.cell(row = 1, column = 2).value == 'Ime' or \
           not sh.cell(row = 1, column = 3).value == 'JMBAG' or \
           not sh.cell(row = 1, column = 4).value == 'Korisničko ime' or \
           not sh.cell(row = 1, column = 5).value == 'Email':
            raise BadWorkbook('Loaded old workbook is not apropriet.')

        if sh.cell(row = 1, column = 6).value == 'Grupa' and \
           sh.cell(row = 1, column = 7).value == 'Oslobođen' and \
           sh.cell(row = 1, column = 8).value == 'Položio X puta' and \
           sh.cell(row = 1, column = 9).value == 'Položio u god.' and \
           sh.cell(row = 1, column = 10).value == 'Ponavljač' and \
           sh.cell(row = 1, column = 12).value == 'Grupa' and \
           sh.cell(row = 1, column = 13).value == 'Dan' and \
           sh.cell(row = 1, column = 14).value == 'Vrijeme' and \
           sh.cell(row = 1, column = 15).value == 'Dvorana':
            return 'old table'
        elif sh.cell(row = 1, column = 6).value == 'Oslobođen' and \
             sh.cell(row = 1, column = 7).value == 'Položio X puta' and \
             sh.cell(row = 1, column = 8).value == 'Položio u god.' and \
             sh.cell(row = 1, column = 9).value == 'Želi ponavljati [+]':
            return 'repeat table'
        else:
            raise BadWorkbook('Loaded old workbook is not apropriet.')
        
    except BadWorkbook as e:
        e.add_note('Prilozena stara excel datoteka za ponavljace nije prikladna!')
        raise
    except Exception:
        Exception.add_note('Unexpected error with old wb validation!')
        raise

# -----------------------------------------------------------
# Loads all data from workbook
# Returns -> tuple[list[Student], list[Group]] = (cours_participants,groups)
# cours_participants: alfabeticly sorted students
# groups: naturaly sorted groups
def LoadInputData(sh: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[Student], list[Group]]:
    num_of_students = sh.max_row - 1
    cours_participants: list[Student] = []
    groups: list[Group] = []
    grouplabels: set[str] = set()
    group: Group = None
    group_zero: Group = Group('G0', 'NaN', 'NaN', 'NaN', 0)  # for students that are not assigned a group

    # get all student and group data
    for student_number in range(num_of_students):
        surname = sh.cell(row = student_number+2, column = 1).value
        name = sh.cell(row = student_number+2, column = 2).value
        email = sh.cell(row = student_number+2, column = 3).value
        jmbag = sh.cell(row = student_number+2, column = 4).value
        
        group_data = sh.cell(row = student_number+2, column = 6).value

        if group_data == 'Još nisu odabrali' or group_data == 'Još nisu svrstani' or group_data == 'Oslobođen':
            group = None
        else:
            # pars string containing group data
            grouplabel, rest = group_data.split(' - ',1)
            day, rest = rest.split(' ', 1)
            time, rest = rest.split(' (')
            room = rest.removesuffix(')')
            time = time.replace(' ', '')

            # create group or update if existing
            if grouplabel in grouplabels:   # update existing group
                group = next((group for group in groups if group.group_label == grouplabel), None)
                group.group_size += 1
            else:   # create new group
                grouplabels.add(grouplabel)
                group = Group(grouplabel,day,time,room,1)
                groups.append(group)

        # create new student
        student = Student(name,surname,email,jmbag)
        if group:   # add group to student if he is in one
            student.group = group
            group.students.append(student)
        else:
            group_zero.group_size+=1
            student.group = group_zero
            group_zero.students.append(student)
        cours_participants.append(student)

    # sort created groups list and students list
    groups.sort(key=lambda x: natural_keys(x.group_label))
    locale.setlocale(locale.LC_COLLATE, 'croatian')
    cours_participants.sort(key=lambda x: locale.strxfrm(x.surname))
    
    if group_zero.group_size > 0:
        groups.append(group_zero)

    for group in groups:
        group.students.sort(key=lambda x: locale.strxfrm(x.surname))

    return cours_participants, groups

# -----------------------------------------------------------
def LoadDataFromOldTable(sh1: openpyxl.worksheet.worksheet.Worksheet, sh2: openpyxl.worksheet.worksheet.Worksheet, cours_participants: list[Student]) -> dict:
    num_of_students = sh1.max_row - 1
    attendance_column = sh2.max_column - 7
    grade_columnn = sh2.max_column - 6
    repeat_students: dict = {}

    repeat_students['old_acad_year'] = sh1.cell(row=1,column=11).value
    for student_number in range(num_of_students):
        jmbag:int = sh1.cell(row = student_number+2, column = 3).value

        if any(student.jmbag == jmbag for student in cours_participants):
            passed_n_times: int = sh1.cell(row = student_number+2, column = 8).value
            passed_in_years: str = sh1.cell(row = student_number+2, column = 9).value
            if not passed_n_times: passed_n_times = 0

            # check if he passed
            if sh2.cell(row=student_number+2, column=attendance_column).value == 'DA' and sh2.cell(row=student_number+2, column=grade_columnn).value >= 0.5:
                passed_last_year = True  
            else:
                passed_last_year = False
    
        if passed_last_year:
            passed_n_times+=1
            if not passed_in_years: passed_in_years = repeat_students['old_acad_year']
            else: passed_in_years = f'{passed_in_years}, {repeat_students['old_acad_year']}'
        if not passed_in_years: passed_in_years = ''
        
        absolved = ''
        if passed_last_year or passed_n_times >= 2: # passed last year or passed min needed times
            absolved = '+'

        repeat_students[jmbag] = [absolved, passed_n_times, passed_in_years]

    return repeat_students

# -----------------------------------------------------------
def LoadDataFromRepeatTable(sh: openpyxl.worksheet.worksheet.Worksheet):
    num_of_students = sh.max_row - 1
    repeat_students: dict = {}

    repeat_students['old_acad_year'] = sh.cell(row=1,column=10).value
    for student_number in range(num_of_students):
        jmbag: int = sh.cell(row = student_number+2, column = 3).value
        absolved: str = ''
        if sh.cell(row = student_number+2, column = 6).value == '+' and not sh.cell(row = student_number+2, column = 9).value == '+':
            absolved = '+'
        passed_n_times: int = sh.cell(row = student_number+2, column = 7).value
        passed_in_years: str = sh.cell(row = student_number+2, column = 8).value
        if not passed_n_times: passed_n_times = 0
        if not passed_in_years: passed_in_years = ''
        repeat_students[jmbag] = [absolved, passed_n_times, passed_in_years]
    return repeat_students

# -----------------------------------------------------------
# Writes sheet with all student data (name, JMABG, username, Email, Group, ...)
def WriteDataSheet(workbook: xlsxwriter.Workbook, worksheet: xlsxwriter.workbook.Worksheet, repeat_students: dict, cours_participants: list[Student], groups: list[Group]):
    format_left = workbook.add_format({'align': 'left'})
    format_center = workbook.add_format({'align': 'center'})
    format_header = workbook.add_format({'font_size': 12, 'bold': False, 'align': 'center', 'bg_color': '#BFBFBF'})
    format_white_text = workbook.add_format({'align': 'center', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})

    worksheet.write('A1', 'Prezime',format_header)
    worksheet.write('B1', 'Ime',format_header)
    worksheet.write('C1', 'JMBAG',format_header)
    worksheet.write('D1', 'Korisničko ime',format_header)
    worksheet.write('E1', 'Email',format_header)
    worksheet.write('F1', 'Grupa',format_header)
    worksheet.write('G1', 'Oslobođen',format_header)
    worksheet.write('H1', 'Položio X puta',format_header)
    worksheet.write('I1', 'Položio u god.',format_header)
    worksheet.write('J1', 'Ponavljač',format_header)
    worksheet.write('K1', f'{settings.acad_year}',format_white_text)

    width1 = len('Prezime')+1
    width2 = len('Ime')+1
    width3 = len('JMBAG')+1
    width4 = len('Korisničko ime')+1
    width5 = len('Email')+1
    width6 = 11 # grupa
    width7 = 15.5 # Oslobođen
    width8 = 19 # Položio X puta
    width9 = 19 # Položio u god.
    width10 = 15 # Ponavljač

    row: int = 2
    for student in cours_participants:
        worksheet.write(f'A{row}', student.surname)
        worksheet.write(f'B{row}', student.name)
        worksheet.write(f'C{row}', student.jmbag)
        worksheet.write(f'D{row}', student.username)
        worksheet.write(f'E{row}', student.email)
        if hasattr(student, 'group'):
            worksheet.write(f'F{row}', f'{student.group.group_label}')
        else:
            worksheet.write(f'F{row}', 'G0')
        
        if len(student.surname) > width1: width1 = len(student.surname)+1
        if len(student.name) > width2: width2 = len(student.name)+1
        if len(str(student.jmbag)) > width3: width3 = len(str(student.jmbag))+4
        if len(student.username) > width4: width4 = len(student.username)+1
        if len(student.email) > width5: width5 = len(student.email)+4
        if hasattr(student, 'group'):
            if len(f'{student.group.group_label}') > width6: width6 = len(f'{student.group.group_label}')+1
        
        if repeat_students:
            if any(jmbag == student.jmbag for jmbag in repeat_students.keys()):
                data: list = repeat_students[student.jmbag]     # [str,int,str] = [absolved, passed_n_times, passed_in_years]
                absolved: str = data[0]
                passed_n_times: int = data[1]
                passed_in_years: str = data[2]
                worksheet.write(f'G{row}', f'{absolved}')
                worksheet.write(f'H{row}', passed_n_times)
                worksheet.write(f'I{row}', f'{passed_in_years}')
                worksheet.write(f'J{row}', '+')


                # data: list = repeat_students[student.jmbag]   # [bool,str,int,str] = [passed, was_absolved, passed_n_times, passed_in_years]
                # if data[0] or data[2]>=2: # passed last year or passed min needed times
                #     worksheet.write(f'G{row}', '+')
                
                # passed_n_times: int = data[2]
                # passed_in_years = data[3]
                # if data[0]:
                #     passed_n_times+=1
                #     if not data[3]: passed_in_years = repeat_students['old_acad_year']
                #     else: passed_in_years = f'{data[3]}, {repeat_students['old_acad_year']}'
                # if not passed_in_years: passed_in_years = ''

                # if len(passed_in_years)>width9: width9 = len(passed_in_years)

                # worksheet.write(f'H{row}', passed_n_times)
                # worksheet.write(f'I{row}', f'{passed_in_years}')
                # worksheet.write(f'J{row}', '+')

        worksheet.set_row(row-1, 18)  # student rows -> height 30px

        row+=1
    
    worksheet.set_column(0, 0, width1, format_left)
    worksheet.set_column(1, 1, width2, format_left)
    worksheet.set_column(2, 2, width3, format_center)
    worksheet.set_column(3, 3, width4, format_center)
    worksheet.set_column(4, 4, width5, format_center)
    worksheet.set_column(5, 5, width6, format_center)
    worksheet.set_column(6, 6, width7, format_center)
    worksheet.set_column(7, 7, width8, format_center)
    worksheet.set_column(8, 8, width9, format_center)
    worksheet.set_column(9, 9, width10, format_center)

    worksheet.write('L1', 'Grupa',format_header)
    worksheet.write('M1', 'Dan',format_header)
    worksheet.write('N1', 'Vrijeme',format_header)
    worksheet.write('O1', 'Dvorana',format_header)

    width1 = len('Grupa')+2
    width2 = len('Dan')+3
    width3 = len('Vrijeme')+2
    width4 = len('Dvorana')+2
    
    row = 2
    for group in groups:
        worksheet.write(f'L{row}', group.group_label)
        worksheet.write(f'M{row}', group.day)
        worksheet.write(f'N{row}', group.time)
        worksheet.write(f'O{row}', group.lab)

        if len(group.group_label) > width1:width1 = len(group.group_label)+1
        if len(group.day) > width2: width2 = len(group.day)+1
        if len(group.time) > width3: width3 = len(group.time)+1
        if len(group.lab) > width4: width4 = len(group.lab)+1
        
        row+=1
        
    worksheet.set_column(11, 11, width1, format_center)
    worksheet.set_column(12, 12, width2, format_center)
    worksheet.set_column(13, 13, width3, format_center)
    worksheet.set_column(14, 14, width4, format_center)

    worksheet.set_row(0, 30)    # set row 1 height

    worksheet.autofilter(0,5, len(cours_participants),9)    # filter by groups and exemptions

# -----------------------------------------------------------
# Writes sheet that containes all points, average and attendance
def WritePointsSheet(workbook: xlsxwriter.Workbook, worksheet: xlsxwriter.workbook.Worksheet, cours_participants: list[Student]):
    format_header = workbook.add_format({'font_size': 12, 'bold': False, 'text_wrap': True, 'align': 'center', 'bg_color': '#BFBFBF', 'left':0, 'right':0, 'border':1, 'bottom':5 , 'top':5})
    format_name_header = workbook.add_format({'font_size': 12, 'bold': False, 'align': 'center', 'bg_color': '#BFBFBF', 'border':1, 'left':5, 'right':5, 'bottom':5 , 'top':5})
    format_average_header = workbook.add_format({'font_size': 12, 'bold': False, 'align': 'center', 'bg_color': '#BFBFBF', 'border':1, 'left':0, 'right':5, 'bottom':5 , 'top':5})
    format_group_header = workbook.add_format({'align': 'center', 'text_wrap': True, 'bg_color': '#BFBFBF'})
    
    format_students = workbook.add_format({'align': 'left', 'border':1, 'left':5, 'right':5, 'bottom':0 , 'top':0})
    format_bottom = workbook.add_format({'border':1, 'left':0, 'right':0, 'bottom':0, 'top':5 })

    format_attendance_cell = workbook.add_format({'align': 'center', 'border':1, 'left':5, 'right':1, 'bottom':0, 'top':0})
    format_grade_cell = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'border':1, 'left':0, 'right':5, 'bottom':0, 'top':0})
    format_point_cell = workbook.add_format({'align': 'center'})

    format_green1_bg = workbook.add_format({'bg_color': '#9BBB59'})
    format_red1_bg = workbook.add_format({'bg_color': '#FF0000'})
    format_green2_bg = workbook.add_format({'bg_color': '#92D050'})
    format_red2_bg = workbook.add_format({'bg_color': '#C00000'})

    worksheet.write('A1', 'Prezime i Ime',format_name_header)
    
    # --------------------------------------
    # write header
    row: int = 0
    col: int = 0
    if settings.using_lab0.get():
        col+=1
        worksheet.write(row, col, 'LAB0[+/-]',format_header)
        worksheet.set_column(col, col, 10)
        first_eval_col = 2  #zero indexed
        ex=0
    elif settings.no_eval_ex0.get():
        col+=1
        worksheet.write(row, col, 'LAB1[+/-]',format_header)
        worksheet.set_column(col, col, 10)
        first_eval_col = 2  #zero indexed
        ex=1
    else:
        first_eval_col = 1  #zero indexed
        ex=0

    while ex < settings.ex_num:
        col+=1
        worksheet.write(row, col, f'LAB{ex+1}',format_header)
        worksheet.set_column(col, col, 10)
        ex+=1

    first_ex_col = 1    #zero indexed
    last_ex_col = col   #zero indexed

    col+=2   # skip one column
    worksheet.write(row, col, f'ZADOVOLJENA PRISUTNOST',format_header)
    worksheet.set_column(col, col, 14)  # col 'ZADOVOLJENA PRISUTNOST' -> width 133px
    attendance_col = col    #zero indexed

    col+=1
    worksheet.write(row, col, f'PROSJEK',format_average_header)
    worksheet.set_column(col, col, 9)   # col 'PROSJEK' -> width 88px
    grade_col = col   #zero indexed

    col+=1
    worksheet.write(row, col, f'GRUPA', format_group_header)
    worksheet.set_column(col, col, 11)
    group_col = col #zero indexed

    col+=3  # skip 2 columns
    worksheet.write(row, col, f'Studenti koji nisu polozili', format_group_header)
    worksheet.write(row, col+1, f'Prikazati    (Da: +, Ne: -)', format_group_header)
    worksheet.write(row, col+2, f'Nije polozilo', format_group_header)
    worksheet.set_column(col, col, 24)
    worksheet.set_column(col+1, col+1, 12)
    worksheet.set_column(col+2, col+2, 12)

    worksheet.set_row(row, 33)    # row 1 -> height 55px

    # -------------------
    # get failed students
    first_student_cell = xl_rowcol_to_cell(1, 0)
    last_student_cell = xl_rowcol_to_cell(len(cours_participants), 0)

    first_attendance_cell = xl_rowcol_to_cell(1, attendance_col)
    last_attendance_cell = xl_rowcol_to_cell(len(cours_participants), attendance_col)
    
    first_grade_cell = xl_rowcol_to_cell(1, grade_col)
    last_grade_cell = xl_rowcol_to_cell(len(cours_participants), grade_col)

    first_failed_cell = xl_rowcol_to_cell(1, col)
    last_failed_cell = xl_rowcol_to_cell(len(cours_participants), col)

    first_absolved_cell = xl_rowcol_to_cell(1, 6)
    last_absolved_cell = xl_rowcol_to_cell(len(cours_participants), 6)
    
    get_students_cell = xl_rowcol_to_cell(1, col+1)
    
    worksheet.write_formula(1,col, f'=IF({get_students_cell}=\"+\",FILTER({first_student_cell}:{last_student_cell},(({first_attendance_cell}:{last_attendance_cell}=\"NE\")+({first_grade_cell}:{last_grade_cell}<1/2))*(Studenti!{first_absolved_cell}:{last_absolved_cell}=\"\")),\"\")')
    worksheet.write(1,col+1, '-', format_point_cell)
    worksheet.write_formula(1,col+2, f'=SUMPRODUCT(({first_failed_cell}:{last_failed_cell}<>\"\")*1)', format_point_cell)

    # -------------------
    # write students
    width1 = len('Prezime i Ime')+1
    for index, student in enumerate(cours_participants):
        row = 1 + index
        worksheet.write(row, 0, student.fullname, format_students)
        for ex in range(last_ex_col+1):
            worksheet.write_blank(row, first_ex_col + ex, 'blank', format_point_cell)
        
        first_ex_cell = xl_rowcol_to_cell(row, first_ex_col)
        first_eval_cell = xl_rowcol_to_cell(row, first_eval_col)
        last_ex_cell = xl_rowcol_to_cell(row, last_ex_col)
        hidden_cell = xl_rowcol_to_cell(row, last_ex_col+1)
        attendance_cell = xl_rowcol_to_cell(row, attendance_col)
        grade_cell = xl_rowcol_to_cell(row, grade_col)
        group_cell = xl_rowcol_to_cell(row, group_col)
        
        if settings.no_eval_ex0.get():
            worksheet.write_formula(hidden_cell, f'=SUM(COUNTIF({first_eval_cell}:{last_ex_cell},\">={int(settings.max_test_points/2)}\"), COUNTIF({first_ex_cell},\"=+\"))>={settings.attendance}')
        else:
            worksheet.write_formula(hidden_cell, f'=COUNTIF({first_ex_cell}:{last_ex_cell},\">={int(settings.max_test_points/2)}\")>={settings.attendance}')
        worksheet.write_formula(attendance_cell, f'=IF({hidden_cell}=TRUE,\"DA\",\"NE\")', format_attendance_cell)
        if settings.no_eval_ex0.get() and not settings.using_lab0.get():
            if settings.not_using_failed_points.get():
                worksheet.write_formula(grade_cell, f'=SUMIF({first_eval_cell}:{last_ex_cell}, \">={int(settings.max_test_points/2)}\")/({settings.ex_num-1}*{settings.max_test_points})', format_grade_cell)
            else:
                worksheet.write_formula(grade_cell, f'=SUM({first_eval_cell}:{last_ex_cell})/({settings.ex_num-1}*{settings.max_test_points})', format_grade_cell)
        else:
            if settings.not_using_failed_points.get():
                worksheet.write_formula(grade_cell, f'=SUMIF({first_eval_cell}:{last_ex_cell}, \">={int(settings.max_test_points/2)}\")/({settings.ex_num}*{settings.max_test_points})', format_grade_cell)
            else:
                worksheet.write_formula(grade_cell, f'=SUM({first_eval_cell}:{last_ex_cell})/({settings.ex_num}*{settings.max_test_points})', format_grade_cell)
        worksheet.write(group_cell, f'{student.group.group_label}', format_point_cell)
        
        if len(student.fullname) > width1: width1 = len(student.fullname)+1
        worksheet.set_row(row, 18)  # student rows -> height 30px

    col=0
    row+=1
    for index in range(group_col):  # does not include group_col
        worksheet.write_blank(row,  col+index, 'blank', format_bottom)
    
    # conditional formating for attendance
    worksheet.conditional_format(1,attendance_col, row-1,attendance_col, {
        'type':     'text',
        'criteria': 'containing',
        'value':    'DA',
        'format':   format_green2_bg
    })
    worksheet.conditional_format(1,attendance_col, row-1,attendance_col, {
        'type':     'text',
        'criteria': 'containing',
        'value':    'NE',
        'format':   format_red2_bg
    })

    # conditional formating for grade
    worksheet.conditional_format(1,grade_col, row-1,grade_col, {
        'type':     'cell',
        'criteria': '>=',
        'value':    settings.min_average_required/100,
        'format':   format_green2_bg
    })
    worksheet.conditional_format(1,grade_col, row-1,grade_col, {
        'type':     'cell',
        'criteria': '<',
        'value':    settings.min_average_required/100,
        'format':   format_red2_bg
    })

    # conditional formating for points
    worksheet.conditional_format(1,first_eval_col, row-1,last_ex_col, {
        'type':     'cell',
        'criteria': 'between',
        'minimum':  settings.max_test_points/2,
        'maximum':  settings.max_test_points,
        'format':   format_green1_bg
    })
    worksheet.conditional_format(1,first_eval_col, row-1,last_ex_col, {
        'type':     'cell',
        'criteria': 'between',
        'minimum':  1,
        'maximum':  settings.max_test_points/2 - 0.001,
        'format':   format_red1_bg
    })

    # conditional formating for no_eval
    if settings.no_eval_ex0.get():
        worksheet.conditional_format(1,first_ex_col, row-1,first_ex_col, {
            'type':     'text',
            'criteria': 'containing',
            'value':    '+',
            'format':   format_green1_bg
        })
        worksheet.conditional_format(1,first_ex_col, row-1,first_ex_col, {
            'type':     'text',
            'criteria': 'containing',
            'value':    '-',
            'format':   format_red1_bg
        })
    
    worksheet.set_column(0,0 ,width1)
    worksheet.autofilter(0,group_col, row-1,group_col)    # filter by groups and exemptions
    worksheet.hide_zero()

# -----------------------------------------------------------
# Writes sheet with all filled group tables
def WriteTablesSheet(workbook: xlsxwriter.Workbook, worksheet: xlsxwriter.workbook.Worksheet, groups: list[Group]):
    format_header = workbook.add_format({'font_size': 15, 'bold': True, 'align': 'center'})
    
    format_fullname_label = workbook.add_format({'border':1, 'left':1, 'right':1, 'bottom':5 , 'top':5})
    format_ex_label = workbook.add_format({'align': 'center', 'bg_color': '#BFBFBF', 'border':1, 'left':1, 'right':1, 'bottom':5 , 'top':5})
    format_last_ex_label = workbook.add_format({'align': 'center', 'bg_color': '#BFBFBF', 'border':1, 'left':1, 'right':5, 'bottom':5 , 'top':5})
    format_group_label = workbook.add_format({'border':1, 'left':5, 'right':5, 'bottom':5 , 'top':5})
    
    format_empty_cell = workbook.add_format({'border':1, 'left':5, 'right':1, 'bottom':1 })
    
    format_first_index_label = workbook.add_format({'align': 'left', 'border':1, 'left':5, 'right':1, 'bottom':1 , 'top':5})
    format_index_label = workbook.add_format({'align': 'left', 'border':1, 'left':5, 'right':1, 'bottom':1 , 'top':1})
    format_last_index_label = workbook.add_format({'align': 'left', 'border':1, 'left':5, 'right':1, 'bottom':5 , 'top':1})

    format_name_cell = workbook.add_format({'border':1, 'right':1, 'bottom':1})
    format_last_name_cell = workbook.add_format({'border':1, 'right':1, 'bottom':5})
    format_center_cell = workbook.add_format({'align': 'center', 'border':1, 'right':1, 'bottom':1})
    format_right_moste_center_cell = workbook.add_format({'align': 'center', 'border':1, 'right':5, 'bottom':1})
    format_last_row_center_cell = workbook.add_format({'align': 'center', 'border':1, 'right':1, 'bottom':5})
    format_right_moste_last_row_center_cell = workbook.add_format({'align': 'center', 'border':1, 'right':5, 'bottom':5})

    format_group0_bottom = workbook.add_format({'border':1, 'left':0, 'right':0, 'bottom':0 , 'top':5})

    if settings.cours_name == '':
        worksheet.write('E1', 'Predmet Smjer - yyyy/yy',format_header)
    else:
        worksheet.write('E1', f'{settings.cours_name} {settings.cours_number} - {settings.acad_year}',format_header)
    worksheet.write('E2', 'LABORATORIJSKE VJEŽBE FESB',format_header)

    max_group_size: int = 0
    for group in groups:
        if group.group_label == 'G0':
            continue
        if max_group_size < group.group_size: max_group_size = group.group_size

    # create ex labels
    ex_label_width = 9
    ex_labels:list[str] = []
    if settings.using_custom_exlabels.get():
        ex_labels = settings.custom_ex_labels
        if settings.no_eval_ex0.get():
            ex_labels[0] = f'{ex_labels[0]}[+/-]'
        for label in ex_labels:
            if len(label)> ex_label_width: ex_label_width = len(label)+2
    else:
        ex = 1
        if settings.using_lab0.get(): 
            ex_labels.append('Lab0[+/0]')
        elif settings.no_eval_ex0.get(): 
            ex_labels.append('Lab1[+/-]')
            ex+=1
        while ex <= settings.ex_num:
            ex_labels.append(f'Lab{ex}')
            ex+=1

    group_label_width = 4
    student_width = len('Prezime i Ime')
    
    gap_between_tables_col = 3  # = wanted col gap + 1
    gap_between_tables_row = 3  # = wanted row gap

    index_col1:int = 0
    first_ex_col1 = index_col1 + 2
    last_ex_col1 = first_ex_col1 + len(ex_labels) - 1

    index_col2 = last_ex_col1 + gap_between_tables_col
    first_ex_col2 = index_col2 + 2
    last_ex_col2 = first_ex_col2 + len(ex_labels) - 1

    index_col_group0 = last_ex_col2 + gap_between_tables_col + 1
    first_ex_col_group0 = index_col_group0 + 2
    last_ex_col_group0 = first_ex_col_group0 + len(ex_labels) - 1

    index_col = index_col1
    table_header_row1: int = 5
    first_ex_col = first_ex_col1
    for counter, group in enumerate(groups):
        table_header_row2 = table_header_row1 + 1
        student_row = table_header_row2 + 1

        first_student_coordinat = {'row':student_row, 'col':index_col+1}
        settings.student_coordinats[group.group_label] = first_student_coordinat
        
        if len(group.group_label) > group_label_width: group_label_width = len(group.group_label)
        worksheet.write(table_header_row1, index_col, f'{group.group_label}', format_group_label)
        worksheet.write(table_header_row1, index_col+1, f'{group.day} {group.time}', format_group_label)
        worksheet.write(table_header_row1, index_col+2, f'{group.lab}', format_group_label)
        worksheet.set_row(table_header_row1, 16.8)  # label rows -> height 28px
        
        worksheet.write_blank(table_header_row2, index_col, 'blank', format_empty_cell)
        worksheet.write(table_header_row2, index_col+1, 'Prezime i Ime', format_fullname_label)
        for ex, ex_label in enumerate(ex_labels):   # write ex labels
            if ex == len(ex_labels)-1: format = format_last_ex_label
            else: format = format_ex_label
            worksheet.write(table_header_row2, first_ex_col+ex, f'{ex_label}', format)
        worksheet.set_row(table_header_row2, 16.8)  # label rows -> height 28px

        index = 1
        for student in group.students:  # add students to table and format row
            if index == 1: format = format_first_index_label
            else: format = format_index_label
            worksheet.write(student_row, index_col, index, format)
            worksheet.write(student_row, index_col+1, f'{student.fullname}', format_name_cell)
            if student_width < len(student.fullname): student_width = len(student.fullname)
            format2 = format_center_cell
            for ex,_ in enumerate(ex_labels):
                if ex == len(ex_labels)-1: format2 = format_right_moste_center_cell
                worksheet.write_blank(student_row, first_ex_col+ex, 'blank', format2)
            worksheet.set_row(student_row, 18)  # student rows -> height 30px
            student_row+=1  # move to next row
            index+=1

        while index < max_group_size+2: # padd with empty group slots
            if index == 1: 
                format = format_first_index_label
                format2 = format_center_cell
                format_name = format_name_cell
            elif index == max_group_size+1: 
                format = format_last_index_label
                format2 = format_last_row_center_cell
                format_name = format_last_name_cell
            else: 
                format = format_index_label
                format2 = format_center_cell
                format_name = format_name_cell
            worksheet.write(student_row, index_col, index, format)
            worksheet.write_blank(student_row, index_col+1, 'blank',format_name)
            for ex,_ in enumerate(ex_labels):
                if index == max_group_size+1 and ex == len(ex_labels)-1: format2 = format_right_moste_last_row_center_cell
                elif ex == len(ex_labels)-1: format2 = format_right_moste_center_cell
                worksheet.write_blank(student_row, first_ex_col+ex, 'blank', format2)
            worksheet.set_row(student_row, 18)  # student rows -> height 30px
            student_row+=1
            index+=1
        
        if group.group_label == 'G0':
            worksheet.write_blank(student_row, index_col, 'blank', format_group0_bottom)
            worksheet.write_blank(student_row, index_col+1, 'blank',format_group0_bottom)
            for ex,_ in enumerate(ex_labels):
                worksheet.write_blank(student_row, first_ex_col+ex, 'blank', format_group0_bottom)

        # move to next table location
        if index_col == 0: 
            index_col = index_col2
            first_ex_col = first_ex_col2
        else:
            table_header_row1 += gap_between_tables_row + max_group_size+1 + 2  # gap + indexed rows + header rows
            index_col = index_col1
            first_ex_col = first_ex_col1
        
        if counter == len(groups)-2 and groups[-1].group_label == 'G0':   # can only true for last group
            index_col = index_col_group0
            table_header_row1 = 2
            first_ex_col = first_ex_col_group0

    worksheet.set_column(index_col1, index_col1, group_label_width)
    worksheet.set_column(index_col2, index_col2, group_label_width)
    worksheet.set_column(index_col_group0, index_col_group0, group_label_width)
    worksheet.set_column(index_col1+1, index_col1+1, student_width)
    worksheet.set_column(index_col2+1, index_col2+1, student_width)
    worksheet.set_column(index_col_group0+1, index_col_group0+1, student_width)
    worksheet.set_column(first_ex_col1, last_ex_col1, ex_label_width)
    worksheet.set_column(first_ex_col2, last_ex_col2, ex_label_width)
    worksheet.set_column(first_ex_col_group0, last_ex_col_group0, ex_label_width)

def WriteScheduleSheet(workbook: xlsxwriter.Workbook, worksheet: xlsxwriter.workbook.Worksheet, groups: list[Group]):
    format_title = workbook.add_format({'font_size': 15, 'bold': True, 'align': 'center'})

    pon_bg_color = '#FF5F1F'
    uto_bg_color = '#1FFF0F'
    sri_bg_color = '#B720F4'
    cet_bg_color = '#CFFF04'
    pet_bg_color = '#B4B4B4'
    err_bg_color = '#000000'

    format_table_header_left_pon = workbook.add_format({'bg_color': pon_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':0})
    format_table_header_center_pon = workbook.add_format({'bg_color': pon_bg_color, 'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':0})
    format_table_header_right_pon = workbook.add_format({'bg_color': pon_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':5})
    
    format_table_header_left_uto = workbook.add_format({'bg_color': uto_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':0})
    format_table_header_center_uto = workbook.add_format({'bg_color': uto_bg_color, 'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':0})
    format_table_header_right_uto = workbook.add_format({'bg_color': uto_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':5})
    
    format_table_header_left_sri = workbook.add_format({'bg_color': sri_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':0})
    format_table_header_center_sri = workbook.add_format({'bg_color': sri_bg_color, 'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':0})
    format_table_header_right_sri = workbook.add_format({'bg_color': sri_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':5})
    
    format_table_header_left_cet = workbook.add_format({'bg_color': cet_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':0})
    format_table_header_center_cet = workbook.add_format({'bg_color': cet_bg_color, 'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':0})
    format_table_header_right_cet = workbook.add_format({'bg_color': cet_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':5})
    
    format_table_header_left_pet = workbook.add_format({'bg_color': pet_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':0})
    format_table_header_center_pet = workbook.add_format({'bg_color': pet_bg_color, 'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':0})
    format_table_header_right_pet = workbook.add_format({'bg_color': pet_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':5})

    format_table_header_left_err = workbook.add_format({'bg_color': err_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':0})
    format_table_header_center_err = workbook.add_format({'bg_color': err_bg_color, 'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':0})
    format_table_header_right_err = workbook.add_format({'bg_color': err_bg_color, 'border':1, 'bottom':5, 'top':5, 'left':0, 'right':5})
    

    format_index = workbook.add_format({'align': 'center', 'border':1, 'bottom':0, 'top':0, 'left':5, 'right':0})
    format_name = workbook.add_format({'align': 'center', 'border':1, 'bottom':5, 'top':5, 'left':5, 'right':5})
    format_empty_right = workbook.add_format({'border':1, 'bottom':0, 'top':0, 'left':0, 'right':5})
    
    format_bottom_left = workbook.add_format({'border':1, 'bottom':5, 'top':0, 'left':5, 'right':0})
    format_bottom_center = workbook.add_format({'border':1, 'bottom':5, 'top':0, 'left':0, 'right':0})
    format_bottom_right = workbook.add_format({'border':1, 'bottom':5, 'top':0, 'left':0, 'right':5})

    if not settings.cours_name: worksheet.write(2,4,f'Predmet Smjer - yyyy/yy', format_title)
    else: worksheet.write(2,4,f'{settings.cours_name} {settings.cours_number} - {settings.acad_year}', format_title)
    group_coordinates = {'row': 4,'col':0}
    row_gap_between_tables = 3
    col_gap_between_tables = 0

    max_group_size: int = 0
    for group in groups:
        if group.group_label == 'G0':
            continue
        if max_group_size < group.group_size: max_group_size = group.group_size

    width = 15  # 110px

    for group in groups:
        if group.group_label == 'G0':
            continue
        starting_row = group_coordinates['row']
        match group.day:
            case 'PON':
                format_table_header_left = format_table_header_left_pon
                format_table_header_center = format_table_header_center_pon
                format_table_header_right = format_table_header_right_pon
            case 'UTO':
                format_table_header_left = format_table_header_left_uto
                format_table_header_center = format_table_header_center_uto
                format_table_header_right = format_table_header_right_uto
            case 'SRI':
                format_table_header_left = format_table_header_left_sri
                format_table_header_center = format_table_header_center_sri
                format_table_header_right = format_table_header_right_sri
            case 'ČET':
                format_table_header_left = format_table_header_left_cet
                format_table_header_center = format_table_header_center_cet
                format_table_header_right = format_table_header_right_cet
            case 'PET':
                format_table_header_left = format_table_header_left_pet
                format_table_header_center = format_table_header_center_pet
                format_table_header_right = format_table_header_right_pet
            case _:
                format_table_header_left = format_table_header_left_err
                format_table_header_center = format_table_header_center_err
                format_table_header_right = format_table_header_right_err
        
        worksheet.write_blank(group_coordinates['row'], group_coordinates['col'], 'blank', format_table_header_left)
        worksheet.write(group_coordinates['row'], group_coordinates['col']+1, f'{group.group_label}', format_table_header_center)
        worksheet.write_blank(group_coordinates['row'], group_coordinates['col']+2, 'blank', format_table_header_right)
        
        group_coordinates['row']+=1
        worksheet.write_blank(group_coordinates['row'], group_coordinates['col'], 'blank', format_table_header_left)
        worksheet.write(group_coordinates['row'], group_coordinates['col']+1, f'Lab {group.lab}', format_table_header_center)
        worksheet.write_blank(group_coordinates['row'], group_coordinates['col']+2, 'blank', format_table_header_right)
        
        group_coordinates['row']+=1
        worksheet.write_blank(group_coordinates['row'], group_coordinates['col'], 'blank', format_table_header_left)
        worksheet.write(group_coordinates['row'], group_coordinates['col']+1, f'{group.day} {group.time}', format_table_header_center)
        worksheet.write_blank(group_coordinates['row'], group_coordinates['col']+2, 'blank', format_table_header_right)
        
        group_coordinates['row']+=1 # empty row
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col'], 'blank', format_index)
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+2, 'blank', format_empty_right)
        group_coordinates['row']+=1 # empty row
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col'], 'blank', format_index)
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+2, 'blank', format_empty_right)

        for index, student in enumerate(group.students):
            group_coordinates['row']+=1
            if len(student.fullname) > width:
                width = len(student.fullname)
            worksheet.write(group_coordinates['row'],group_coordinates['col'], index+1, format_index)
            worksheet.write(group_coordinates['row'],group_coordinates['col']+1, f'{student.fullname}', format_name)
            worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+2, 'blank', format_empty_right)

        while index < max_group_size-1:
            index+=1
            group_coordinates['row']+=1
            worksheet.write(group_coordinates['row'],group_coordinates['col'], index+1, format_index)
            worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+1, 'blank', format_name)
            worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+2, 'blank', format_empty_right)

        group_coordinates['row']+=1 # empty row
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col'], 'blank', format_index)
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+2, 'blank', format_empty_right)

        group_coordinates['row']+=1 # empty row
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col'], 'blank', format_bottom_left)
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+1, 'blank', format_bottom_center)
        worksheet.write_blank(group_coordinates['row'],group_coordinates['col']+2, 'blank', format_bottom_right)

        dif = group_coordinates['row'] - starting_row
        group_coordinates['row'] = starting_row
        
        if group_coordinates['col'] == 0 or group_coordinates['col'] == 3:
            group_coordinates['col'] += 3 + col_gap_between_tables
        else:
            group_coordinates['col'] = 0
            group_coordinates['row'] += dif + row_gap_between_tables
    
    worksheet.set_column(0, 0, 5)
    worksheet.set_column(1, 1, width)
    worksheet.set_column(2, 2, 5)
    worksheet.set_column(3, 3, 5)
    worksheet.set_column(4, 4, width)
    worksheet.set_column(5, 5, 5)
    worksheet.set_column(6, 6, 5)
    worksheet.set_column(7, 7, width)
    worksheet.set_column(8, 8, 5)
    
# -----------------------------------------------------------
def LinkTableAndPointsSheet(workbook: xlsxwriter.Workbook, worksheet: xlsxwriter.workbook.Worksheet, cours_participants: list[Student]):
    format_point_cell = workbook.add_format({'align': 'center'})
    
    if settings.using_lab0.get(): total_ex_num = settings.ex_num+1
    else: total_ex_num = settings.ex_num
    for index, student in enumerate(cours_participants):
        # if student.group.group_label == 'G0':
        #     continue
        student_table_coordinates = settings.student_coordinats[student.group.group_label]
        student_name = xl_rowcol_to_cell(student_table_coordinates['row'], student_table_coordinates['col'], True,True)
        last_ex = xl_rowcol_to_cell(student_table_coordinates['row'], student_table_coordinates['col']+total_ex_num, True,True)
        row = index+1
        index_array_possition = 1
        for ex in range(total_ex_num):
            col = ex+1
            index_array_possition+=1
            worksheet.write_formula(row,col, f'=@INDEX(Tablice!{student_name}:{last_ex},,{index_array_possition})', format_point_cell)

        settings.student_coordinats[student.group.group_label]['row'] += 1

    worksheet.set_column(col+1,col+1, None, None,{'hidden': True})
    
# -----------------------------------------------------------
# util functions
def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    return [ atoi(c) for c in re.split(r'(\d+)', text) ]

# -----------------------------------------------------------
# Main function for call
# Returns -> bool = (success)
def gen_tables(input_file: str, old_file:str = None)-> tuple[bool,str]:
    logger = logging.getLogger('my_app.table_gen.exel')
    logger.setLevel('INFO')
    try:
        new_file = 'data/cours_workbook.xlsx'

        input_wb = openpyxl.load_workbook(filename=input_file)
        input_sh = input_wb.worksheets[0]

        msg = CheckForValidInputWorkbook(input_wb,input_sh)
        logger.info(f'Valid wb from {msg}.')

        cours_participants, groups = LoadInputData(input_sh)
        logger.info('Loaded input data from input wb.')

        if old_file:
            old_wb = openpyxl.load_workbook(filename=old_file, data_only=True)
            old_sh1 = old_wb.worksheets[0]
            old_wb_type = CheckForValidOldWorkbook(old_wb, old_sh1)
            logger.info(f'Valid {old_wb_type} wb.')            
            old_sh1 = old_wb['Studenti']
            if old_wb_type == 'old table':
                old_sh2 = old_wb['Bodovi']
                repeat_students = LoadDataFromOldTable(old_sh1, old_sh2, cours_participants)
                logger.info('Loaded old data from old wb.')
            elif old_wb_type == 'repeat table':
                repeat_students = LoadDataFromRepeatTable(old_sh1)
                logger.info('Loaded data from repeat students wb.')
                pass
        else: repeat_students = None

        out_wb = xlsxwriter.Workbook(new_file)
        data_sheet = out_wb.add_worksheet('Studenti')
        point_worksheet = out_wb.add_worksheet('Bodovi')
        table_worksheet = out_wb.add_worksheet('Tablice')
        schedule_worksheet = out_wb.add_worksheet('Raspored')
        logger.info('Created output wb and sheets.')
        
        WritePointsSheet(out_wb, point_worksheet, cours_participants)
        WriteTablesSheet(out_wb, table_worksheet, groups)
        LinkTableAndPointsSheet(out_wb, point_worksheet, cours_participants)
        WriteDataSheet(out_wb, data_sheet, repeat_students, cours_participants, groups)
        WriteScheduleSheet(out_wb, schedule_worksheet, groups)
        logger.info('Filled all sheets in wb.')

        out_wb.close()

        CopyAndRename(srcname='cours_workbook.xlsx', dstname='lab_tablice')
        logger.info('Copied wb to downloads folder.')
        
        delpath=Path(new_file)
        delpath.unlink()
    
    except BadWorkbook:
        raise
    except Exception as e:
        raise