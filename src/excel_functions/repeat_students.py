from labgenpackage.classes import Student
from gui.util import CopyAndRename
from pathlib import Path

import openpyxl.worksheet.worksheet
import openpyxl.worksheet
import openpyxl
import gui.settings as settings
import xlsxwriter, locale, logging

# custom Exception class
class BadWorkbook(ValueError):
    pass

def CheckForValidOldWorkbook(wb: openpyxl.Workbook, sh: openpyxl.worksheet.worksheet.Worksheet):
    try:
        if not len(wb.sheetnames) == 4 or \
           not sh.max_column == 15 or \
           not sh.cell(row = 1, column = 1).value == 'Prezime' or \
           not sh.cell(row = 1, column = 2).value == 'Ime' or \
           not sh.cell(row = 1, column = 3).value == 'JMBAG' or \
           not sh.cell(row = 1, column = 4).value == 'Korisničko ime' or \
           not sh.cell(row = 1, column = 5).value == 'Email' or \
           not sh.cell(row = 1, column = 6).value == 'Grupa' or \
           not sh.cell(row = 1, column = 7).value == 'Oslobođen' or \
           not sh.cell(row = 1, column = 8).value == 'Položio X puta' or \
           not sh.cell(row = 1, column = 9).value == 'Položio u god.' or \
           not sh.cell(row = 1, column = 10).value == 'Ponavljač' or \
           not sh.cell(row = 1, column = 12).value == 'Grupa' or \
           not sh.cell(row = 1, column = 13).value == 'Dan' or \
           not sh.cell(row = 1, column = 14).value == 'Vrijeme' or \
           not sh.cell(row = 1, column = 15).value == 'Dvorana':
            raise BadWorkbook('Loaded old workbook is not apropriet.')
        else:
            pass
        
    except BadWorkbook as e:
        e.add_note('Prilozena stara excel datoteka za ponavljace nije prikladna!')
        raise
    except Exception:
        Exception.add_note('Unexpected error with old wb validation!')
        raise


def CheckForValidExemptWorkbook(wb: openpyxl.Workbook, sh: openpyxl.worksheet.worksheet.Worksheet):
    try:
        if not len(wb.sheetnames) == 1 or \
           not sh.max_column == 10 or \
           not sh.cell(row = 1, column = 1).value == 'Prezime' or \
           not sh.cell(row = 1, column = 2).value == 'Ime' or \
           not sh.cell(row = 1, column = 3).value == 'JMBAG' or \
           not sh.cell(row = 1, column = 4).value == 'Korisničko ime' or \
           not sh.cell(row = 1, column = 5).value == 'Email' or \
           not sh.cell(row = 1, column = 6).value == 'Oslobođen' or \
           not sh.cell(row = 1, column = 7).value == 'Položio X puta' or \
           not sh.cell(row = 1, column = 8).value == 'Položio u god.' or \
           not sh.cell(row = 1, column = 9).value == 'Želi ponavljati [+]':
            raise BadWorkbook('Loaded old workbook is not apropriet.')
        else:
            pass
        
    except BadWorkbook as e:
        e.add_note('Prilozena stara excel datoteka za ponavljace nije prikladna!')
        raise
    except Exception:
        Exception.add_note('Unexpected error with old wb validation!')
        raise


def LoadExemptData(sh: openpyxl.worksheet.worksheet.Worksheet) -> list[int]:
    num_of_students = sh.max_row - 1
    exempt_students: list[int] = []
    for student_number in range(num_of_students):
        if sh.cell(row=student_number+2, column=6).value == '+' and not sh.cell(row=student_number+2, column=9).value == '+':
            jmbag = sh.cell(row=student_number+2, column=3).value
            exempt_students.append(jmbag)
    return exempt_students


def LoadOldData(sh1: openpyxl.worksheet.worksheet.Worksheet, sh2: openpyxl.worksheet.worksheet.Worksheet, cours_participants: list[Student]) -> dict:
    num_of_students = sh1.max_row - 1
    attendance_column = sh2.max_column - 7
    grade_columnn = sh2.max_column - 6
    repeat_students: dict = {}

    repeat_students['old_acad_year'] = sh1.cell(row=1,column=11).value
    for student_number in range(num_of_students):
        jmbag:int = sh1.cell(row = student_number+2, column = 3).value

        if any(student.jmbag == jmbag for student in cours_participants):
            passed_n_times: int = sh1.cell(row = student_number+2, column = 8).value
            passed_in_years: str = sh1.cell(row = student_number+2, column = 9).value
            if not passed_n_times: passed_n_times = 0

            # check if he passed
            if sh2.cell(row=student_number+2, column=attendance_column).value == 'DA' and sh2.cell(row=student_number+2, column=grade_columnn).value >= 0.5:
                repeat_students[jmbag] = [True, passed_n_times, passed_in_years]
            else:
                repeat_students[jmbag] = [False, passed_n_times, passed_in_years]
    return repeat_students

def WriteDataSheet(workbook: xlsxwriter.Workbook, worksheet: xlsxwriter.workbook.Worksheet, repeat_students: dict, cours_participants: list[Student]):
    format_left = workbook.add_format({'align': 'left'})
    format_center = workbook.add_format({'align': 'center'})
    format_header = workbook.add_format({'font_size': 12, 'bold': False, 'align': 'center', 'bg_color': '#BFBFBF'})
    format_white_text = workbook.add_format({'align': 'center', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})

    format_jmbag = workbook.add_format({'num_format': '0000000000', 'align': 'center'})

    worksheet.write('A1', 'Prezime',format_header)
    worksheet.write('B1', 'Ime',format_header)
    worksheet.write('C1', 'JMBAG',format_header)
    worksheet.write('D1', 'Korisničko ime',format_header)
    worksheet.write('E1', 'Email',format_header)
    
    worksheet.write('F1', 'Oslobođen',format_header)
    worksheet.write('G1', 'Položio X puta',format_header)
    worksheet.write('H1', 'Položio u god.',format_header)
    worksheet.write('I1', 'Želi ponavljati [+]',format_header)
    worksheet.write('J1', f'{settings.acad_year}',format_white_text)

    width1 = len('Prezime')+1
    width2 = len('Ime')+1
    width3 = len('JMBAG')+1
    width4 = len('Korisničko ime')+1
    width5 = len('Email')+1
    width6 = 15.5 # Oslobođen
    width7 = 19 # Položio X puta
    width8 = 19 # Položio u god.
    width9 = 21.5 # Zeli ponavljati

    row: int = 2
    if not repeat_students:
        return
    
    for student in cours_participants:
        if any(jmbag == student.jmbag for jmbag in repeat_students.keys()):
            worksheet.write(f'A{row}', student.surname)
            worksheet.write(f'B{row}', student.name)
            worksheet.write(f'C{row}', student.jmbag, format_jmbag)
            worksheet.write(f'D{row}', student.username)
            worksheet.write(f'E{row}', student.email)

            if len(student.surname) > width1: width1 = len(student.surname)+1
            if len(student.name) > width2: width2 = len(student.name)+1
            if len(str(student.jmbag)) > width3: width3 = len(str(student.jmbag))+4
            if len(student.username) > width4: width4 = len(student.username)+1
            if len(student.email) > width5: width5 = len(student.email)+4

            data: list = repeat_students[student.jmbag]   # [bool,str,int,str] = [passed, passed_n_times, passed_in_years]
            passed_last:bool = data[0]
            passed_n_times:int = data[1]
            passed_in_years = data[2]
            
            if passed_last:
                passed_n_times+=1
                if not passed_in_years: passed_in_years = repeat_students['old_acad_year']
                else: passed_in_years = f'{passed_in_years}, {repeat_students['old_acad_year']}'
            
            if not passed_in_years: passed_in_years = ''
            if len(passed_in_years)>width8: width8 = len(passed_in_years)

            if (settings.freed_if_passed_last.get() and passed_last) or passed_n_times >= settings.num_of_needed_passes:
                worksheet.write(f'F{row}', '+')
            
            worksheet.write(f'G{row}', passed_n_times)
            worksheet.write(f'H{row}', f'{passed_in_years}')

            worksheet.set_row(row-1, 18)  # student rows -> height 30px

            row+=1
    
    worksheet.set_column(0, 0, width1, format_left)
    worksheet.set_column(1, 1, width2, format_left)
    worksheet.set_column(2, 2, width3, format_center)
    worksheet.set_column(3, 3, width4, format_center)
    worksheet.set_column(4, 4, width5, format_center)
    worksheet.set_column(5, 5, width6, format_center)
    worksheet.set_column(6, 6, width7, format_center)
    worksheet.set_column(7, 7, width8, format_center)
    worksheet.set_column(8, 8, width9, format_center)

    worksheet.set_row(0, 30)    # set row 1 height

    worksheet.autofilter(0,5, len(repeat_students),8)    # filter by groups and exemptions


def gen_repeat_students(old_file:str = None, cours_participants:dict[str, Student] = None):
    logger = logging.getLogger('my_app.repeat_gen.exel')
    logger.setLevel('INFO')
    try:
        new_file = 'data/repeats_workbook.xlsx'

        cours_participants_list = list(cours_participants.values())
        locale.setlocale(locale.LC_COLLATE, 'croatian')
        cours_participants_list.sort(key=lambda x: locale.strxfrm(x.surname))

        if old_file:
            old_wb = openpyxl.load_workbook(filename=old_file, data_only=True)
            old_sh1 = old_wb.worksheets[0]
            CheckForValidOldWorkbook(old_wb, old_sh1)
            logger.info('Valid old wb.')
            
            old_sh1 = old_wb['Studenti']
            old_sh2 = old_wb['Bodovi']
            repeat_students = LoadOldData(old_sh1, old_sh2, cours_participants_list)
            logger.info('Loaded old data from old wb.')
        else: repeat_students = None

        out_wb = xlsxwriter.Workbook(new_file)
        data_sheet = out_wb.add_worksheet('Studenti')

        logger.info('Created output wb and sheets.')
        
        WriteDataSheet(out_wb, data_sheet, repeat_students, cours_participants_list)

        logger.info('Filled all sheets in wb.')

        out_wb.close()

        CopyAndRename(srcpath='data/repeats_workbook.xlsx', dstname='ponavljaci')
        logger.info('Copied wb to downloads folder.')
        
        delpath=Path(new_file)
        delpath.unlink()
    
    except BadWorkbook:
        raise
    except Exception as e:
        raise


def get_exempt_students(exempt_student_file: str) -> list[int]:
    logger = logging.getLogger('my_app.repeat_gen.exel')
    logger.setLevel('INFO')
    try:
        exempt_wb = openpyxl.load_workbook(filename=exempt_student_file, data_only=True)
        exempt_sh = exempt_wb['Studenti']

        CheckForValidExemptWorkbook(exempt_wb, exempt_sh)
        logger.info('Valid old wb.')

        exempt_students = LoadExemptData(exempt_sh)
        logger.info('Loaded old data from old wb.')
        return exempt_students
    
    except BadWorkbook as e:
        raise ValueError from e
    except Exception as e:
        raise